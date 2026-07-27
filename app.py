# -*- coding: utf-8 -*-
"""
项目调试实施管理库
-------------------
一个轻量的 Web 应用，用于管理多个项目从「立项/开始 → 调试 → 实施 → 结束」
的全生命周期，并记录项目结束后的「售后故障配合」。

技术栈：Flask + SQLite（零外部依赖，开箱即用）。
运行：  python3 app.py   然后浏览器访问 http://127.0.0.1:5000
"""

import os
import sqlite3
from datetime import datetime, date

from flask import (
    Flask, request, jsonify, render_template,
    redirect, url_for, flash, send_file, session
)
from io import BytesIO
import json as _json

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get("DB_PATH", os.path.join(BASE_DIR, "pm.db"))

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "pm-lib-secret-key-2024")
# 模板改动后自动重载，避免手动重启才能生效（管理库为内部工具，开销可忽略）
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.jinja_env.auto_reload = True

# 硬编码管理员账号（内部工具用，后续可改为数据库用户表）
ADMIN_USER = "admin"
ADMIN_PASS = "mrd-admin"


def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

# ---------------------------------------------------------------------------
# 字典/选项（统一在后端定义，前端直接引用，避免散落各处）
# ---------------------------------------------------------------------------
PROJECT_STATUS = ["未开始", "设计联络阶段", "调试中", "实施中", "已结束", "售后中"]
TASK_CATEGORY = ["调试", "实施", "排故", "其他"]
TASK_STATUS = ["待开始", "进行中", "已完成", "阻塞"]
TASK_SUPPORT_MODE = ["远程支持", "出差现场支持", "外出现场支持"]
FAULT_SEVERITY = ["紧急", "严重", "一般"]
FAULT_STATUS = ["待处理", "处理中", "已解决"]
FAULT_RESPONSE_MODE = ["远程响应", "出差现场响应", "外出现场响应"]
PLATFORM_OPTIONS = ["上海总部", "广州平台", "北京平台", "南京平台"]


# ---------------------------------------------------------------------------
# 数据库
# ---------------------------------------------------------------------------
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    # 默认关闭,必须显式开启,ON DELETE CASCADE 才能生效
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    conn = get_db()
    # 建表连接也显式开启外键约束
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.cursor()
    cur.executescript(
        """
        CREATE TABLE IF NOT EXISTS projects (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT NOT NULL,
            customer      TEXT,
            manager       TEXT,
            platform      TEXT,
            status        TEXT NOT NULL DEFAULT '未开始',
            start_date    TEXT,
            plan_end_date TEXT,
            actual_end_date TEXT,
            description   TEXT,
            created_at    TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS tasks (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id    INTEGER NOT NULL,
            title         TEXT NOT NULL,
            category      TEXT DEFAULT '实施',
            assignee      TEXT,
            start_date    TEXT,
            plan_end_date TEXT,
            actual_end_date TEXT,
            support_mode  TEXT DEFAULT '远程支持',
            work_hours    REAL DEFAULT 0,
            status        TEXT DEFAULT '待开始',
            note          TEXT,
            created_at    TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS faults (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id    INTEGER NOT NULL,
            title         TEXT NOT NULL,
            severity      TEXT DEFAULT '一般',
            status        TEXT DEFAULT '待处理',
            handler       TEXT,
            reported_date TEXT,
            response_date TEXT,
            response_mode TEXT DEFAULT '远程响应',
            resolved_date TEXT,
            work_hours    REAL DEFAULT 0,
            note          TEXT,
            created_at    TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS devices (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id  INTEGER NOT NULL,
            model       TEXT,
            quantity    INTEGER DEFAULT 0,
            note        TEXT,
            created_at  TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS members (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS project_members (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            member_id  INTEGER NOT NULL,
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE,
            FOREIGN KEY (member_id) REFERENCES members(id) ON DELETE CASCADE
        );
        """
    )
    conn.commit()
    conn.close()


# 实施人员初始名单（仅在人员表为空时写入，保证可扩展）
DEFAULT_MEMBERS = ["金恺", "张振武", "张小强", "班志鹏", "吴智伟", "王嘉旗",
                  "吕文敬", "付博文", "陈威", "许思腾", "万宇辉", "范秋良", "罗小增"]


def migrate_db():
    """对已有库做轻量迁移。"""
    conn = get_db()
    cur = conn.cursor()
    # devices 表：移除 brand / spec 列（如存在）
    cols = [r[1] for r in cur.execute("PRAGMA table_info(devices)").fetchall()]
    if "brand" in cols or "spec" in cols:
        cur.executescript(
            """
            CREATE TABLE devices_new (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id  INTEGER NOT NULL,
                model       TEXT,
                quantity    INTEGER DEFAULT 0,
                note        TEXT,
                created_at  TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
            );
            INSERT INTO devices_new (id, project_id, model, quantity, note, created_at)
                SELECT id, project_id, model, quantity, note, created_at FROM devices;
            DROP TABLE devices;
            ALTER TABLE devices_new RENAME TO devices;
            """
        )
        conn.commit()
    # projects 表：新增 platform 列（如不存在）
    pcols = [r[1] for r in cur.execute("PRAGMA table_info(projects)").fetchall()]
    if "platform" not in pcols:
        cur.execute("ALTER TABLE projects ADD COLUMN platform TEXT")
        conn.commit()
    # tasks 表：重命名字段 due_date→plan_end_date，新增 start_date / actual_end_date（如存在旧结构）
    tcols = [r[1] for r in cur.execute("PRAGMA table_info(tasks)").fetchall()]
    if "due_date" in tcols and "plan_end_date" not in tcols:
        cur.execute("ALTER TABLE tasks RENAME COLUMN due_date TO plan_end_date")
        conn.commit()
        tcols = [r[1] for r in cur.execute("PRAGMA table_info(tasks)").fetchall()]
    if "start_date" not in tcols:
        cur.execute("ALTER TABLE tasks ADD COLUMN start_date TEXT")
        conn.commit()
    if "actual_end_date" not in tcols:
        cur.execute("ALTER TABLE tasks ADD COLUMN actual_end_date TEXT")
        conn.commit()
    # 删除 completed_date 列（如存在），数据已迁移至 actual_end_date
    tcols2 = [r[1] for r in cur.execute("PRAGMA table_info(tasks)").fetchall()]
    if "completed_date" in tcols2:
        # SQLite 不能直接 DROP COLUMN（3.35+ 支持），采用重建方式
        cur.executescript(
            """
            CREATE TABLE tasks_new (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id    INTEGER NOT NULL,
                title         TEXT NOT NULL,
                category      TEXT DEFAULT '实施',
                assignee      TEXT,
                start_date    TEXT,
                plan_end_date TEXT,
                actual_end_date TEXT,
                status        TEXT DEFAULT '待开始',
                note          TEXT,
                created_at    TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
            );
            INSERT INTO tasks_new (id, project_id, title, category, assignee,
                                   start_date, plan_end_date, actual_end_date, status, note, created_at)
                SELECT id, project_id, title, category, assignee,
                       NULL, plan_end_date, completed_date, status, note, created_at
                FROM tasks;
            DROP TABLE tasks;
            ALTER TABLE tasks_new RENAME TO tasks;
            """
        )
        conn.commit()
    # tasks 表：新增 support_mode 列（如不存在）
    tcols3 = [r[1] for r in cur.execute("PRAGMA table_info(tasks)").fetchall()]
    if "support_mode" not in tcols3:
        cur.execute("ALTER TABLE tasks ADD COLUMN support_mode TEXT DEFAULT '远程支持'")
        conn.commit()
    if "work_hours" not in tcols3:
        cur.execute("ALTER TABLE tasks ADD COLUMN work_hours REAL DEFAULT 0")
        conn.commit()
    # faults 表：新增 work_hours 列（如不存在）
    fcols = [r[1] for r in cur.execute("PRAGMA table_info(faults)").fetchall()]
    if "work_hours" not in fcols:
        cur.execute("ALTER TABLE faults ADD COLUMN work_hours REAL DEFAULT 0")
        conn.commit()
    if "response_date" not in fcols:
        cur.execute("ALTER TABLE faults ADD COLUMN response_date TEXT")
        conn.commit()
    if "response_mode" not in fcols:
        cur.execute("ALTER TABLE faults ADD COLUMN response_mode TEXT DEFAULT '远程响应'")
        conn.commit()
    # members 表：不存在则建；为空则写入初始名单
    cur.execute(
        """CREATE TABLE IF NOT EXISTS members (
               id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE)"""
    )
    cur.execute(
        """CREATE TABLE IF NOT EXISTS project_members (
               id INTEGER PRIMARY KEY AUTOINCREMENT,
               project_id INTEGER NOT NULL,
               member_id INTEGER NOT NULL,
               FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE,
               FOREIGN KEY (member_id) REFERENCES members(id) ON DELETE CASCADE)"""
    )
    if cur.execute("SELECT COUNT(*) FROM members").fetchone()[0] == 0:
        cur.executemany("INSERT INTO members (name) VALUES (?)",
                        [(n,) for n in DEFAULT_MEMBERS])
        conn.commit()
    conn.close()


def seed_db():
    """写入示例数据，方便首次体验。仅当项目表为空时写入。"""
    conn = get_db()
    cur = conn.cursor()
    cnt = cur.execute("SELECT COUNT(*) AS c FROM projects").fetchone()["c"]
    if cnt > 0:
        conn.close()
        return

    today = date.today().isoformat()
    cur.execute(
        """INSERT INTO projects
           (name, customer, manager, platform, status, start_date, plan_end_date, description)
           VALUES (?,?,?,?,?,?,?,?)""",
        ("A 工厂产线自动化改造", "某装备集团", "张工", "上海总部", "实施中",
         today, "", "整线 PLC 与视觉检测系统集成。"),
    )
    pid1 = cur.lastrowid
    cur.execute(
        """INSERT INTO projects
           (name, customer, manager, platform, status, start_date, plan_end_date, actual_end_date, description)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        ("B 楼宇温控系统", "某地产公司", "李工", "广州平台", "售后中",
         today, today, today, "已完成调试交付，进入售后保障期。"),
    )
    pid2 = cur.lastrowid

    tasks = [
        (pid1, "现场设备接线与通电", "实施", "王工", today, today, "", "远程支持", 2.5, "进行中", "配电柜就位"),
        (pid1, "PLC 程序下载与联调", "调试", "张工", "", today, "", "出差现场支持", 0, "待开始", ""),
        (pid2, "系统功能验收", "调试", "李工", today, today, today, "外出现场支持", 1.0, "已完成", "一次性通过"),
    ]
    for t in tasks:
        cur.execute(
            """INSERT INTO tasks
               (project_id, title, category, assignee, start_date, plan_end_date, actual_end_date, support_mode, work_hours, status, note)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            t,
        )

    faults = [
        (pid2, "3 楼温度传感器偶发掉线", "一般", "待处理", "李工", today, "", "远程响应", "", 1.0, "需现场复现"),
        (pid2, "控制柜风扇异响", "严重", "处理中", "赵工", today, today, "出差现场响应", "", 0.5, "已下单更换"),
    ]
    for f in faults:
        cur.execute(
            """INSERT INTO faults
               (project_id, title, severity, status, handler, reported_date, response_date, response_mode, resolved_date, work_hours, note)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            f,
        )

    devices = [
        (pid1, "PLC 控制器", 4, "主控制模块"),
        (pid1, "工业相机", 8, "视觉检测工位"),
        (pid2, "温度传感器", 12, "楼栋各层部署"),
    ]
    for d in devices:
        cur.execute(
            """INSERT INTO devices
               (project_id, model, quantity, note)
               VALUES (?,?,?,?)""",
            d,
        )
    conn.commit()
    conn.close()


# ---------------------------------------------------------------------------
# 工具
# ---------------------------------------------------------------------------
def today_str():
    return date.today().isoformat()


def fmt_date(s):
    """Jinja 模板用的日期格式化：空值显示破折号。"""
    return s if s else "—"


app.jinja_env.globals["fmtDate"] = fmt_date


def rows_to_dicts(rows):
    """sqlite3.Row 不能直接 JSON 序列化，转换为 dict 列表。"""
    return [dict(r) for r in rows]


# ---------------------------------------------------------------------------
# 登录 / 登出
# ---------------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        user = (request.form.get("username") or "").strip()
        pwd = request.form.get("password") or ""
        if user == ADMIN_USER and pwd == ADMIN_PASS:
            session["user"] = user
            return redirect(url_for("dashboard"))
        error = "用户名或密码错误"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login"))


# ---------------------------------------------------------------------------
# 页面路由（服务端渲染）
# ---------------------------------------------------------------------------
@app.route("/")
@login_required
def dashboard():
    platform = request.args.get("platform", "")
    conn = get_db()
    cur = conn.cursor()

    proj_total = cur.execute(
        "SELECT COUNT(*) c FROM projects" + (" WHERE platform = ?" if platform else ""),
        (platform,) if platform else ()
    ).fetchone()["c"]
    proj_by_status = cur.execute(
        "SELECT status, COUNT(*) c FROM projects" + (" WHERE platform = ?" if platform else "") + " GROUP BY status",
        (platform,) if platform else ()
    ).fetchall()

    fault_total = cur.execute(
        "SELECT COUNT(*) c FROM faults f JOIN projects p ON f.project_id = p.id" + (" WHERE p.platform = ?" if platform else ""),
        (platform,) if platform else ()
    ).fetchone()["c"]
    fault_by_status = cur.execute(
        "SELECT f.status, COUNT(*) c FROM faults f JOIN projects p ON f.project_id = p.id" + (" WHERE p.platform = ?" if platform else "") + " GROUP BY f.status",
        (platform,) if platform else ()
    ).fetchall()
    fault_open = cur.execute(
        "SELECT COUNT(*) c FROM faults f JOIN projects p ON f.project_id = p.id WHERE f.status != '已解决'" + (" AND p.platform = ?" if platform else ""),
        (platform,) if platform else ()
    ).fetchone()["c"]

    blocked_tasks = cur.execute(
        "SELECT COUNT(*) c FROM tasks t JOIN projects p ON t.project_id = p.id WHERE t.status = '阻塞'" + (" AND p.platform = ?" if platform else ""),
        (platform,) if platform else ()
    ).fetchone()["c"]

    # 即将到期（未完成的任务，按计划完成日期升序取前 5）
    upcoming = cur.execute(
        """SELECT t.*, p.name AS project_name
           FROM tasks t JOIN projects p ON t.project_id = p.id
           WHERE t.status != '已完成' AND t.plan_end_date IS NOT NULL""" + (" AND p.platform = ?" if platform else "") + """
           ORDER BY t.plan_end_date ASC LIMIT 5""",
        (platform,) if platform else ()
    ).fetchall()

    # 未解决故障（按严重度）
    open_faults = cur.execute(
        """SELECT f.*, p.name AS project_name
           FROM faults f JOIN projects p ON f.project_id = p.id
           WHERE f.status != '已解决'""" + (" AND p.platform = ?" if platform else "") + """
           ORDER BY CASE f.severity WHEN '紧急' THEN 0 WHEN '严重' THEN 1 ELSE 2 END, f.reported_date ASC""",
        (platform,) if platform else ()
    ).fetchall()

    # 待完成的调试/实施任务（所有项目，未完成状态）
    pending_tasks = cur.execute(
        """SELECT t.*, p.name AS project_name
           FROM tasks t JOIN projects p ON t.project_id = p.id
           WHERE t.status != '已完成'""" + (" AND p.platform = ?" if platform else "") + """
           ORDER BY t.plan_end_date ASC""",
        (platform,) if platform else ()
    ).fetchall()

    conn.close()

    # 构建 JSON 供前端柱状图使用
    proj_status_map = {r["status"]: r["c"] for r in proj_by_status}
    fault_status_map = {r["status"]: r["c"] for r in fault_by_status}

    return render_template(
        "dashboard.html",
        proj_total=proj_total,
        proj_by_status=proj_by_status,
        proj_by_status_json=_json.dumps(proj_status_map, ensure_ascii=False),
        fault_total=fault_total,
        fault_by_status=fault_by_status,
        fault_by_status_json=_json.dumps(fault_status_map, ensure_ascii=False),
        fault_open=fault_open,
        blocked_tasks=blocked_tasks,
        upcoming=upcoming,
        open_faults=open_faults,
        pending_tasks=pending_tasks,
        PROJECT_STATUS=PROJECT_STATUS,
        PLATFORM_OPTIONS=PLATFORM_OPTIONS,
        current_platform=platform,
    )


@app.route("/api/dashboard")
@login_required
def api_dashboard():
    platform = request.args.get("platform", "")
    conn = get_db()
    cur = conn.cursor()
    proj_total = cur.execute(
        "SELECT COUNT(*) c FROM projects" + (" WHERE platform = ?" if platform else ""),
        (platform,) if platform else ()
    ).fetchone()["c"]
    proj_by_status = [dict(r) for r in cur.execute(
        "SELECT status, COUNT(*) c FROM projects" + (" WHERE platform = ?" if platform else "") + " GROUP BY status",
        (platform,) if platform else ()
    ).fetchall()]
    fault_total = cur.execute(
        "SELECT COUNT(*) c FROM faults f JOIN projects p ON f.project_id = p.id" + (" WHERE p.platform = ?" if platform else ""),
        (platform,) if platform else ()
    ).fetchone()["c"]
    fault_open = cur.execute(
        "SELECT COUNT(*) c FROM faults f JOIN projects p ON f.project_id = p.id WHERE f.status != '已解决'" + (" AND p.platform = ?" if platform else ""),
        (platform,) if platform else ()
    ).fetchone()["c"]
    blocked = cur.execute(
        "SELECT COUNT(*) c FROM tasks t JOIN projects p ON t.project_id = p.id WHERE t.status = '阻塞'" + (" AND p.platform = ?" if platform else ""),
        (platform,) if platform else ()
    ).fetchone()["c"]
    upcoming = [dict(r) for r in cur.execute(
        """SELECT t.*, p.name AS project_name
           FROM tasks t JOIN projects p ON t.project_id = p.id
           WHERE t.status != '已完成' AND t.plan_end_date IS NOT NULL""" + (" AND p.platform = ?" if platform else "") + """
           ORDER BY t.plan_end_date ASC LIMIT 5""",
        (platform,) if platform else ()
    ).fetchall()]
    open_faults = [dict(r) for r in cur.execute(
        """SELECT f.*, p.name AS project_name
           FROM faults f JOIN projects p ON f.project_id = p.id
           WHERE f.status != '已解决'""" + (" AND p.platform = ?" if platform else "") + """
           ORDER BY CASE f.severity WHEN '紧急' THEN 0 WHEN '严重' THEN 1 ELSE 2 END, f.reported_date ASC""",
        (platform,) if platform else ()
    ).fetchall()]
    pending_tasks = [dict(r) for r in cur.execute(
        """SELECT t.*, p.name AS project_name
           FROM tasks t JOIN projects p ON t.project_id = p.id
           WHERE t.status != '已完成'""" + (" AND p.platform = ?" if platform else "") + """
           ORDER BY t.plan_end_date ASC""",
        (platform,) if platform else ()
    ).fetchall()]
    conn.close()
    return jsonify(ok=True, proj_total=proj_total, proj_by_status=proj_by_status,
                   fault_total=fault_total, fault_open=fault_open, blocked_tasks=blocked,
                   upcoming=upcoming, open_faults=open_faults, pending_tasks=pending_tasks)


@app.route("/projects")
@login_required
def projects():
    status = request.args.get("status", "")
    platform = request.args.get("platform", "")
    kw = request.args.get("q", "").strip()
    conn = get_db()
    cur = conn.cursor()
    sql = "SELECT * FROM projects WHERE 1=1"
    args = []
    if status:
        sql += " AND status = ?"
        args.append(status)
    if platform:
        sql += " AND platform = ?"
        args.append(platform)
    if kw:
        sql += " AND (name LIKE ? OR customer LIKE ? OR manager LIKE ?)"
        args += [f"%{kw}%"] * 3
    sql += " ORDER BY created_at DESC"
    rows = rows_to_dicts(cur.execute(sql, args).fetchall())
    conn.close()
    return render_template(
        "projects.html",
        projects=rows,
        status=status,
        platform=platform,
        q=kw,
        PROJECT_STATUS=PROJECT_STATUS,
        PLATFORM_OPTIONS=PLATFORM_OPTIONS,
    )


@app.route("/project/<int:pid>")
@login_required
def project_detail(pid):
    conn = get_db()
    cur = conn.cursor()
    proj = cur.execute("SELECT * FROM projects WHERE id = ?", (pid,)).fetchone()
    if not proj:
        conn.close()
        flash("项目不存在")
        return redirect(url_for("projects"))
    tasks = rows_to_dicts(cur.execute(
        "SELECT * FROM tasks WHERE project_id = ? ORDER BY plan_end_date ASC", (pid,)
    ).fetchall())
    faults = rows_to_dicts(cur.execute(
        "SELECT * FROM faults WHERE project_id = ? ORDER BY reported_date DESC", (pid,)
    ).fetchall())
    devices = rows_to_dicts(cur.execute(
        "SELECT * FROM devices WHERE project_id = ? ORDER BY id ASC", (pid,)
    ).fetchall())
    member_rows = cur.execute(
        """SELECT m.id AS mid, m.name FROM project_members pm
           JOIN members m ON pm.member_id = m.id WHERE pm.project_id = ? ORDER BY m.id""",
        (pid,),
    ).fetchall()
    proj_member_ids = [r["mid"] for r in member_rows]
    proj_members = [r["name"] for r in member_rows]
    members_all = [dict(r) for r in cur.execute("SELECT id, name FROM members ORDER BY id")]

    # 工时统计（任务 + 故障合并），支持日期范围筛选
    date_start = request.args.get("date_start", "")
    date_end = request.args.get("date_end", "")
    task_date_where = ""
    fault_date_where = ""
    date_args = []
    if date_start:
        task_date_where += " AND (start_date >= ? OR plan_end_date >= ?)"
        fault_date_where += " AND (reported_date >= ? OR resolved_date >= ?)"
        date_args = [date_start, date_start, date_start, date_start]
    if date_end:
        task_date_where += " AND (start_date <= ? OR plan_end_date <= ?)"
        fault_date_where += " AND (reported_date <= ? OR resolved_date <= ?)"
        date_args += [date_end, date_end, date_end, date_end]

    task_hours_total = cur.execute(
        "SELECT COALESCE(SUM(work_hours), 0) FROM tasks WHERE project_id = ?" + task_date_where,
        (pid, *date_args[:2] if date_start else *(), *date_args[2:4] if date_end else ())
    ).fetchone()[0]

    task_hours_remote = cur.execute(
        "SELECT COALESCE(SUM(work_hours), 0) FROM tasks WHERE project_id = ? AND support_mode = '远程支持'" + task_date_where,
        (pid, *date_args[:2] if date_start else *(), *date_args[2:4] if date_end else ())
    ).fetchone()[0]

    task_hours_onsite = cur.execute(
        "SELECT COALESCE(SUM(work_hours), 0) FROM tasks WHERE project_id = ? AND support_mode = '出差现场支持'" + task_date_where,
        (pid, *date_args[:2] if date_start else *(), *date_args[2:4] if date_end else ())
    ).fetchone()[0]

    task_hours_outside = cur.execute(
        "SELECT COALESCE(SUM(work_hours), 0) FROM tasks WHERE project_id = ? AND support_mode = '外出现场支持'" + task_date_where,
        (pid, *date_args[:2] if date_start else *(), *date_args[2:4] if date_end else ())
    ).fetchone()[0]

    # 故障工时日期过滤用 reported_date 和 resolved_date
    fds = [date_start, date_start] if date_start else []
    fde = [date_end, date_end] if date_end else []
    f_args = tuple(fds + fde)

    fault_hours_total = cur.execute(
        "SELECT COALESCE(SUM(work_hours), 0) FROM faults WHERE project_id = ?" + fault_date_where,
        (pid, *f_args)
    ).fetchone()[0]

    fault_hours_remote = cur.execute(
        "SELECT COALESCE(SUM(work_hours), 0) FROM faults WHERE project_id = ? AND response_mode = '远程响应'" + fault_date_where,
        (pid, *f_args)
    ).fetchone()[0]

    fault_hours_onsite = cur.execute(
        "SELECT COALESCE(SUM(work_hours), 0) FROM faults WHERE project_id = ? AND response_mode = '出差现场响应'" + fault_date_where,
        (pid, *f_args)
    ).fetchone()[0]

    fault_hours_outside = cur.execute(
        "SELECT COALESCE(SUM(work_hours), 0) FROM faults WHERE project_id = ? AND response_mode = '外出现场响应'" + fault_date_where,
        (pid, *f_args)
    ).fetchone()[0]

    # 合并总工时
    hours_total = task_hours_total + fault_hours_total
    hours_remote = task_hours_remote + fault_hours_remote
    hours_onsite = task_hours_onsite + fault_hours_onsite
    hours_outside = task_hours_outside + fault_hours_outside

    conn.close()
    return render_template(
        "project_detail.html",
        proj=proj,
        proj_member_ids=proj_member_ids,
        proj_members=proj_members,
        tasks=tasks,
        faults=faults,
        devices=devices,
        hours_total=hours_total,
        hours_remote=hours_remote,
        hours_onsite=hours_onsite,
        hours_outside=hours_outside,
        task_hours_total=task_hours_total,
        fault_hours_total=fault_hours_total,
        fault_hours_remote=fault_hours_remote,
        fault_hours_onsite=fault_hours_onsite,
        fault_hours_outside=fault_hours_outside,
        date_start=date_start,
        date_end=date_end,
        TASK_CATEGORY=TASK_CATEGORY,
        TASK_STATUS=TASK_STATUS,
        TASK_SUPPORT_MODE=TASK_SUPPORT_MODE,
        FAULT_SEVERITY=FAULT_SEVERITY,
        FAULT_STATUS=FAULT_STATUS,
        FAULT_RESPONSE_MODE=FAULT_RESPONSE_MODE,
        PLATFORM_OPTIONS=PLATFORM_OPTIONS,
        MEMBERS=members_all,
        today=today_str(),
    )


@app.route("/faults")
@login_required
def faults_view():
    status = request.args.get("status", "")
    platform = request.args.get("platform", "")
    conn = get_db()
    cur = conn.cursor()
    sql = """SELECT f.*, p.name AS project_name, p.platform AS platform FROM faults f
             JOIN projects p ON f.project_id = p.id WHERE 1=1"""
    args = []
    if status:
        sql += " AND f.status = ?"
        args.append(status)
    if platform:
        sql += " AND p.platform = ?"
        args.append(platform)
    sql += " ORDER BY CASE f.severity WHEN '紧急' THEN 0 WHEN '严重' THEN 1 ELSE 2 END, f.reported_date DESC"
    rows = cur.execute(sql, args).fetchall()
    conn.close()
    return render_template(
        "faults.html",
        faults=rows,
        status=status,
        platform=platform,
        FAULT_STATUS=FAULT_STATUS,
        FAULT_SEVERITY=FAULT_SEVERITY,
        PLATFORM_OPTIONS=PLATFORM_OPTIONS,
    )


def _save_members(cur, pid, members):
    """替换某项目的实施人员关联（members 为成员 id 列表）。"""
    cur.execute("DELETE FROM project_members WHERE project_id = ?", (pid,))
    if not members:
        return
    for mid in members:
        try:
            mid = int(mid)
        except (TypeError, ValueError):
            continue
        cur.execute(
            "INSERT OR IGNORE INTO project_members (project_id, member_id) VALUES (?, ?)",
            (pid, mid),
        )


# ---------------------------------------------------------------------------
# API：项目 CRUD
# ---------------------------------------------------------------------------
@app.route("/api/projects", methods=["POST"])
def api_project_create():
    d = request.json
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify(ok=False, msg="项目名称不能为空"), 400
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """INSERT INTO projects
           (name, customer, manager, platform, status, start_date, plan_end_date, actual_end_date, description)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (
            name,
            d.get("customer", ""),
            d.get("manager", ""),
            d.get("platform", "") or None,
            d.get("status", "未开始"),
            d.get("start_date", "") or None,
            d.get("plan_end_date", "") or None,
            d.get("actual_end_date", "") or None,
            d.get("description", ""),
        ),
    )
    conn.commit()
    new_id = cur.lastrowid
    _save_members(cur, new_id, d.get("members"))
    conn.commit()
    conn.close()
    return jsonify(ok=True, id=new_id)


@app.route("/api/projects/<int:pid>", methods=["PUT"])
def api_project_update(pid):
    d = request.json
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id FROM projects WHERE id = ?", (pid,))
    if not cur.fetchone():
        conn.close()
        return jsonify(ok=False, msg="项目不存在"), 404
    cur.execute(
        """UPDATE projects SET
           name=?, customer=?, manager=?, platform=?, status=?,
           start_date=?, plan_end_date=?, actual_end_date=?, description=?
           WHERE id=?""",
        (
            d.get("name", ""),
            d.get("customer", ""),
            d.get("manager", ""),
            d.get("platform", "") or None,
            d.get("status", "未开始"),
            d.get("start_date", "") or None,
            d.get("plan_end_date", "") or None,
            d.get("actual_end_date", "") or None,
            d.get("description", ""),
            pid,
        ),
    )
    _save_members(cur, pid, d.get("members"))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/projects/<int:pid>", methods=["DELETE"])
def api_project_delete(pid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM projects WHERE id = ?", (pid,))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ---------------------------------------------------------------------------
# API：任务 CRUD
# ---------------------------------------------------------------------------
@app.route("/api/tasks", methods=["POST"])
def api_task_create():
    d = request.json
    pid = d.get("project_id")
    title = (d.get("title") or "").strip()
    if not pid or not title:
        return jsonify(ok=False, msg="缺少项目或任务标题"), 400
    actual_end_date = d.get("actual_end_date", "") or None
    if d.get("status") == "已完成" and not actual_end_date:
        actual_end_date = today_str()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """INSERT INTO tasks
           (project_id, title, category, assignee, start_date, plan_end_date, actual_end_date, support_mode, work_hours, status, note)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (
            pid, title, d.get("category", "实施"),
            d.get("assignee", ""), d.get("start_date", "") or None,
            d.get("plan_end_date", "") or None, d.get("actual_end_date", "") or None,
            d.get("support_mode", "远程支持"), d.get("work_hours", 0) or 0,
            d.get("status", "待开始"), d.get("note", ""),
        ),
    )
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/tasks/<int:tid>", methods=["PUT"])
def api_task_update(tid):
    d = request.json
    actual_end_date = d.get("actual_end_date", "") or None
    if d.get("status") == "已完成" and not actual_end_date:
        actual_end_date = today_str()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """UPDATE tasks SET
           title=?, category=?, assignee=?, start_date=?, plan_end_date=?, actual_end_date=?,
           support_mode=?, work_hours=?, status=?, note=?
           WHERE id=?""",
        (
            d.get("title", ""), d.get("category", "实施"),
            d.get("assignee", ""), d.get("start_date", "") or None,
            d.get("plan_end_date", "") or None, actual_end_date,
            d.get("support_mode", "远程支持"), d.get("work_hours", 0) or 0,
            d.get("status", "待开始"), d.get("note", ""), tid,
        ),
    )
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/tasks/<int:tid>", methods=["DELETE"])
def api_task_delete(tid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM tasks WHERE id = ?", (tid,))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ---------------------------------------------------------------------------
# API：故障 CRUD
# ---------------------------------------------------------------------------
@app.route("/api/faults", methods=["POST"])
def api_fault_create():
    d = request.json
    pid = d.get("project_id")
    title = (d.get("title") or "").strip()
    if not pid or not title:
        return jsonify(ok=False, msg="缺少项目或故障描述"), 400
    resolved_date = d.get("resolved_date", "") or None
    if d.get("status") == "已解决" and not resolved_date:
        resolved_date = today_str()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """INSERT INTO faults
           (project_id, title, severity, status, handler, reported_date, response_date, response_mode, resolved_date, work_hours, note)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (
            pid, title, d.get("severity", "一般"), d.get("status", "待处理"),
            d.get("handler", ""), d.get("reported_date", "") or today_str(),
            d.get("response_date", "") or None, d.get("response_mode", "远程响应"),
            resolved_date, d.get("work_hours", 0) or 0, d.get("note", ""),
        ),
    )
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/faults/<int:fid>", methods=["PUT"])
def api_fault_update(fid):
    d = request.json
    resolved_date = d.get("resolved_date", "") or None
    if d.get("status") == "已解决" and not resolved_date:
        resolved_date = today_str()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """UPDATE faults SET
           title=?, severity=?, status=?, handler=?, reported_date=?, response_date=?, response_mode=?, resolved_date=?, work_hours=?, note=?
           WHERE id=?""",
        (
            d.get("title", ""), d.get("severity", "一般"),
            d.get("status", "待处理"), d.get("handler", ""),
            d.get("reported_date", "") or today_str(),
            d.get("response_date", "") or None, d.get("response_mode", "远程响应"),
            resolved_date, d.get("work_hours", 0) or 0, d.get("note", ""), fid,
        ),
    )
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/faults/<int:fid>", methods=["DELETE"])
def api_fault_delete(fid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM faults WHERE id = ?", (fid,))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ---------------------------------------------------------------------------
# API：设备 CRUD
# ---------------------------------------------------------------------------
@app.route("/api/devices", methods=["POST"])
def api_device_create():
    d = request.json
    pid = d.get("project_id")
    model = (d.get("model") or "").strip()
    if not pid or not model:
        return jsonify(ok=False, msg="缺少项目或设备型号"), 400
    try:
        qty = int(d.get("quantity") or 0)
    except (ValueError, TypeError):
        qty = 0
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """INSERT INTO devices (project_id, model, quantity, note)
           VALUES (?,?,?,?)""",
        (pid, model, qty, d.get("note", "").strip()),
    )
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/devices/<int:did>", methods=["PUT"])
def api_device_update(did):
    d = request.json
    try:
        qty = int(d.get("quantity") or 0)
    except (ValueError, TypeError):
        qty = 0
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """UPDATE devices SET model=?, quantity=?, note=? WHERE id=?""",
        (d.get("model", "").strip(), qty, d.get("note", "").strip(), did),
    )
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@app.route("/api/devices/<int:did>", methods=["DELETE"])
def api_device_delete(did):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM devices WHERE id = ?", (did,))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ---------------------------------------------------------------------------
# API：实施人员（成员）管理
# ---------------------------------------------------------------------------
@app.route("/api/members", methods=["GET"])
def api_member_list():
    conn = get_db()
    rows = conn.execute("SELECT id, name FROM members ORDER BY id").fetchall()
    conn.close()
    return jsonify(ok=True, members=[dict(r) for r in rows])


@app.route("/api/members", methods=["POST"])
def api_member_create():
    name = (request.json or {}).get("name", "").strip()
    if not name:
        return jsonify(ok=False, msg="人员姓名不能为空"), 400
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO members (name) VALUES (?)", (name,))
        new_id = cur.lastrowid
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify(ok=False, msg="该人员已存在"), 409
    conn.close()
    return jsonify(ok=True, id=new_id, name=name)


@app.route("/api/members/<int:mid>", methods=["DELETE"])
def api_member_delete(mid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM members WHERE id = ?", (mid,))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ---------------------------------------------------------------------------
# 数据导入 / 模板
# ---------------------------------------------------------------------------
def _cell(v):
    """把 NaN / 空值归一为 None。"""
    if v is None:
        return None
    try:
        if isinstance(v, float) and pd_isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, str) and v.strip() == "":
        return None
    return v


def _date(v):
    if v is None:
        return None
    try:
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
    except Exception:
        pass
    s = str(v).strip()
    if s in ("", "NaT", "nan", "None", "NaT"):
        return None
    return s


def _num(v):
    """把单元格值转为数字，失败返回 0。"""
    if v is None:
        return 0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0


@app.route("/api/import-template")
def api_import_template():
    from openpyxl import Workbook
    wb = Workbook()
    sheets = {
        "项目": ["项目名称", "客户", "负责人", "所属平台", "状态", "开始日期", "计划结束", "实际结束", "描述"],
        "任务": ["项目名称", "任务标题", "类型", "参与人员", "开始日期", "计划完成日期", "实际完成日期", "支持方式", "工时(人/天)", "状态", "备注"],
        "故障": ["项目名称", "故障描述", "级别", "状态", "处理人", "报告日期", "响应日期", "响应方式", "解决日期", "工时(人/天)", "备注"],
        "设备": ["项目名称", "设备型号", "数量", "备注"],
    }
    for name, headers in sheets.items():
        wb.create_sheet(title=name).append(headers)
    # 项目表放一行示例，方便对照
    wb["项目"].append(["示例项目A", "某客户", "张三", "上海总部", "实施中",
                       "2026-01-01", "2026-06-30", "", "整线调试"])
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name="项目导入模板.xlsx", as_attachment=True)


@app.route("/api/import", methods=["POST"])
def api_import():
    f = request.files.get("file")
    if not f:
        return jsonify(ok=False, msg="未收到文件"), 400
    try:
        import pandas as pd
        xls = pd.read_excel(BytesIO(f.read()), sheet_name=None)
    except Exception as e:
        return jsonify(ok=False, msg="文件解析失败：" + str(e)), 400

    counts = {"projects": 0, "tasks": 0, "faults": 0, "devices": 0}
    conn = get_db()
    cur = conn.cursor()
    name_to_id = {}

    # 项目
    if "项目" in xls:
        for _, r in xls["项目"].fillna("").iterrows():
            name = str(_cell(r.get("项目名称")) or "").strip()
            if not name:
                continue
            platform = str(_cell(r.get("所属平台")) or "").strip() or None
            if platform and platform not in PLATFORM_OPTIONS:
                platform = None
            status = str(_cell(r.get("状态")) or "").strip() or "未开始"
            if status not in PROJECT_STATUS:
                status = "未开始"
            cur.execute(
                """INSERT INTO projects
                   (name, customer, manager, platform, status, start_date, plan_end_date, actual_end_date, description)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (name, str(_cell(r.get("客户")) or "").strip() or None,
                 str(_cell(r.get("负责人")) or "").strip() or None,
                 platform, status, _date(_cell(r.get("开始日期"))),
                 _date(_cell(r.get("计划结束"))), _date(_cell(r.get("实际结束"))),
                 str(_cell(r.get("描述")) or "").strip() or None),
            )
            name_to_id[name] = cur.lastrowid
            counts["projects"] += 1

    # 任务
    if "任务" in xls:
        for _, r in xls["任务"].fillna("").iterrows():
            pid = name_to_id.get(str(_cell(r.get("项目名称")) or "").strip())
            if not pid:
                continue
            title = str(_cell(r.get("任务标题")) or "").strip()
            if not title:
                continue
            category = str(_cell(r.get("类型")) or "").strip() or "实施"
            if category not in TASK_CATEGORY:
                category = "实施"
            status = str(_cell(r.get("状态")) or "").strip() or "待开始"
            if status not in TASK_STATUS:
                status = "待开始"
            cur.execute(
                """INSERT INTO tasks
                   (project_id, title, category, assignee, start_date, plan_end_date, actual_end_date, support_mode, work_hours, status, note)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (pid, title, category, str(_cell(r.get("参与人员")) or "").strip() or None,
                 _date(_cell(r.get("开始日期"))), _date(_cell(r.get("计划完成日期"))),
                 _date(_cell(r.get("实际完成日期"))) if status == "已完成" else None,
                 str(_cell(r.get("支持方式")) or "").strip() or "远程支持",
                 _num(_cell(r.get("工时(人/天)"))),
                 status, str(_cell(r.get("备注")) or "").strip() or None),
            )
            counts["tasks"] += 1

    # 故障
    if "故障" in xls:
        for _, r in xls["故障"].fillna("").iterrows():
            pid = name_to_id.get(str(_cell(r.get("项目名称")) or "").strip())
            if not pid:
                continue
            title = str(_cell(r.get("故障描述")) or "").strip()
            if not title:
                continue
            severity = str(_cell(r.get("级别")) or "").strip() or "一般"
            if severity not in FAULT_SEVERITY:
                severity = "一般"
            status = str(_cell(r.get("状态")) or "").strip() or "待处理"
            if status not in FAULT_STATUS:
                status = "待处理"
            resolved = _date(_cell(r.get("解决日期"))) if status == "已解决" else None
            cur.execute(
                """INSERT INTO faults
                   (project_id, title, severity, status, handler, reported_date, response_date, response_mode, resolved_date, work_hours, note)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (pid, title, severity, status,
                 str(_cell(r.get("处理人")) or "").strip() or None,
                 _date(_cell(r.get("报告日期"))) or today_str(),
                 _date(_cell(r.get("响应日期"))),
                 str(_cell(r.get("响应方式")) or "").strip() or "远程响应",
                 resolved, _num(_cell(r.get("工时(人/天)"))),
                 str(_cell(r.get("备注")) or "").strip() or None),
            )
            counts["faults"] += 1

    # 设备
    if "设备" in xls:
        for _, r in xls["设备"].fillna("").iterrows():
            pid = name_to_id.get(str(_cell(r.get("项目名称")) or "").strip())
            if not pid:
                continue
            model = str(_cell(r.get("设备型号")) or "").strip()
            if not model:
                continue
            q = _cell(r.get("数量"))
            try:
                qty = 0 if (q is None or pd_isna(q)) else int(float(q))
            except Exception:
                qty = 0
            cur.execute(
                "INSERT INTO devices (project_id, model, quantity, note) VALUES (?,?,?,?)",
                (pid, model, qty, str(_cell(r.get("备注")) or "").strip() or None),
            )
            counts["devices"] += 1

    conn.commit()
    conn.close()
    return jsonify(ok=True, counts=counts)


@app.route("/api/export")
def api_export():
    from openpyxl import Workbook
    conn = get_db()
    cur = conn.cursor()
    projects = cur.execute("SELECT * FROM projects ORDER BY id").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "项目"
    ws.append(["项目名称", "客户", "负责人", "所属平台", "状态", "开始日期", "计划结束", "实际结束", "描述"])
    tws = wb.create_sheet("任务")
    tws.append(["项目名称", "任务标题", "类型", "参与人员", "开始日期", "计划完成日期", "实际完成日期", "支持方式", "工时(人/天)", "状态", "备注"])
    fws = wb.create_sheet("故障")
    fws.append(["项目名称", "故障描述", "级别", "状态", "处理人", "报告日期", "响应日期", "响应方式", "解决日期", "工时(人/天)", "备注"])
    dws = wb.create_sheet("设备")
    dws.append(["项目名称", "设备型号", "数量", "备注"])

    for p in projects:
        ws.append([p["name"], p["customer"], p["manager"], p["platform"], p["status"],
                   p["start_date"], p["plan_end_date"], p["actual_end_date"], p["description"]])
        for t in cur.execute("SELECT * FROM tasks WHERE project_id=? ORDER BY id", (p["id"],)):
            tws.append([p["name"], t["title"], t["category"], t["assignee"],
                        t["start_date"], t["plan_end_date"], t["actual_end_date"],
                        t["support_mode"], t["work_hours"], t["status"], t["note"]])
        for f in cur.execute("SELECT * FROM faults WHERE project_id=? ORDER BY id", (p["id"],)):
            fws.append([p["name"], f["title"], f["severity"], f["status"],
                        f["handler"], f["reported_date"], f["response_date"],
                        f["response_mode"], f["resolved_date"], f["work_hours"], f["note"]])
        for d in cur.execute("SELECT * FROM devices WHERE project_id=? ORDER BY id", (p["id"],)):
            dws.append([p["name"], d["model"], d["quantity"], d["note"]])

    conn.close()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name="项目调试实施管理库_导出.xlsx", as_attachment=True)


# ---------------------------------------------------------------------------
# 启动
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    # 确保静态文件目录存在
    os.makedirs(os.path.join(BASE_DIR, "static"), exist_ok=True)
    init_db()
    migrate_db()
    seed_db()
    # 非调试模式：稳定托管（supervisord / 后台运行），不启用 reloader
    app.run(host="0.0.0.0", port=5000, debug=False)
